"""Reporte de diferencias + export Excel 1:1 con round-trip (tarea A9).

El round-trip es la prueba cruzada: el archivo que exporta la Persona 1 tiene que
ser reconocido por el `data/ingest.py` de la Persona 2. Cargamos ese módulo tal
cual y le pasamos el export por su propio detector de columnas.
"""

import importlib.util
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.reportes.diferencias import FilaCierre, calcular_diferencias
from app.reportes.export_excel import ENCABEZADOS, construir_libro

RAIZ = Path(__file__).resolve().parents[3]  # backend/tests/pipeline → raíz del repo


def _cargar_ingest_de_p2():
    ruta = RAIZ / "data" / "ingest.py"
    spec = importlib.util.spec_from_file_location("ingest_p2", ruta)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


FILAS = [
    FilaCierre(articulo_id=1, nr_articulo=95026919, articulo="CAZUELA 16 ONZ",
               unidad="Unidad", sd_teorico=10.0, cantidad_fisica=10.0, orden_original=2),
    FilaCierre(articulo_id=2, nr_articulo=95004459, articulo="CINTA SELLAMIENTO",
               unidad="Unidad", sd_teorico=14.0, cantidad_fisica=17.0, orden_original=1),
    FilaCierre(articulo_id=3, nr_articulo=None, articulo="AGUA 280 ML",
               unidad="Unidad", sd_teorico=10.0, cantidad_fisica=90.0, orden_original=3),
]


def test_semaforo():
    rep = calcular_diferencias(3, "almacen general", FILAS + [
        FilaCierre(articulo_id=4, nr_articulo=7, articulo="X", unidad="Unidad",
                   sd_teorico=10.0, cantidad_fisica=10.0, orden_original=4,
                   anomalia_confirmada=True),
    ])
    por_nombre = {d.articulo: d for d in rep.filas}
    assert por_nombre["CAZUELA 16 ONZ"].semaforo == "normal"      # 0 %
    assert por_nombre["CINTA SELLAMIENTO"].semaforo == "revisar"  # ~21 %
    assert por_nombre["AGUA 280 ML"].semaforo == "critico"        # 800 %
    assert por_nombre["X"].semaforo == "critico"                  # anomalía confirmada
    assert rep.resumen.normal == 1 and rep.resumen.revisar == 1 and rep.resumen.critico == 2
    assert rep.filas[0].semaforo == "critico"  # los críticos van primero


def test_export_estructura_1a1(tmp_path):
    wb = construir_libro({"almacen general": FILAS})
    ruta = tmp_path / "cierre.xlsx"
    wb.save(ruta)

    libro = load_workbook(ruta)
    assert libro.sheetnames == ["almacen general"]
    hoja = libro["almacen general"]
    encabezados = [c.value for c in hoja[1]]
    assert encabezados == ENCABEZADOS  # CANTIDAD, Nr.Artículo, Artículo, Unidad, SD

    # Filas en orden_original; SD = cantidad física; CANTIDAD = consecutivo.
    filas = list(hoja.iter_rows(min_row=2, values_only=True))
    assert filas[0] == (1, 95004459, "CINTA SELLAMIENTO", "Unidad", 17.0)
    assert filas[1] == (2, 95026919, "CAZUELA 16 ONZ", "Unidad", 10.0)
    assert filas[2][0] == 3 and filas[2][1] is None  # artículo sin código


def test_roundtrip_ingest_de_p2_reconoce_el_export(tmp_path):
    """Exportar → el detector de columnas de la ingesta de P2 lo acepta."""
    ingest = _cargar_ingest_de_p2()
    ruta = tmp_path / "cierre.xlsx"
    construir_libro({"almacen general": FILAS}).save(ruta)

    df = pd.read_excel(ruta, sheet_name="almacen general")
    cols = ingest._columnas(df)
    assert cols is not None  # la ingesta reconoce las 5 columnas del insumo
    assert set(cols) == {"nr_articulo", "cantidad", "articulo", "unidad", "sd"}
    # y los valores del round-trip cuadran con lo exportado
    fila_cinta = df[df["Artículo"] == "CINTA SELLAMIENTO"].iloc[0]
    assert fila_cinta[cols["sd"]] == 17.0
    assert fila_cinta[cols["cantidad"]] == 1


def test_titulo_hoja_saneado():
    wb = construir_libro({"rest/fonda [ayb]: piscilago con nombre larguísimo x": []})
    titulo = wb.sheetnames[0]
    assert len(titulo) <= 31
    assert not any(c in titulo for c in r"\/*?:[]")


@pytest.mark.integration
def test_filas_cierre_consulta_bd():
    """Contra Postgres real (I3). Sin DATABASE_URL se omite."""
    import os

    url = os.environ.get("DATABASE_URL")
    if not url:
        pytest.skip("DATABASE_URL no configurada")
    from sqlalchemy import create_engine

    from app.reportes.consultas import filas_cierre

    nombre, filas = filas_cierre(create_engine(url), 3)
    assert isinstance(nombre, str)
    assert all(f.cantidad_fisica is not None for f in filas)
